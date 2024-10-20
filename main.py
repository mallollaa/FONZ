import pandas as pd
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT_DIR = 'Inputs'
KAM_FILE = os.path.join(INPUT_DIR, 'KAM_NAMES.xlsx')
OUTPUT_DIR = '/Users/manalalajmi/PycharmProjects/FonzBOT/FonzOUTPUT'

current_date = datetime.now()
output_file_name = f'Fonz_Commission_Report_{current_date.strftime("%B_%Y")}.xlsx'
output_file_path = os.path.join(OUTPUT_DIR, output_file_name)

os.makedirs(OUTPUT_DIR, exist_ok=True)


def calculate_slab(package_price):
    if 0 <= package_price <= 6:
        return 1.75
    elif 6 < package_price <= 15:
        return 2
    elif package_price > 15:
        return 2.5
    return 0


def format_accounting(value):
    """
    Format a number in accounting style: (xxx) for negative numbers.
    """
    if value < 0:
        return f"({abs(value):,.2f})"
    else:
        return f"{value:,.2f}"


def process_churn(event_df):
    event_df = event_df.dropna(axis=1, how='all')
    event_df['six months completed'] = event_df['Age'].apply(
        lambda x: 'YES' if x > 180 else 'NO'
    )

    event_df = event_df[event_df['six months completed'] == 'NO']
    event_df['Package Price'] = event_df['Amount'] + event_df['DiscountAmount']
    event_df['Slab'] = event_df['Package Price'].apply(calculate_slab)
    event_df['Commission (KD)'] = -1 * event_df['Package Price'] * event_df['Slab']
    total_commission = event_df['Commission (KD)'].sum()
    formatted_total = format_accounting(total_commission)
    total_row = pd.DataFrame([{col: "" for col in event_df.columns},
                              {'Commission (KD)': formatted_total}], index=['', 'Total'])
    event_df = pd.concat([event_df, total_row], ignore_index=True)

    return event_df
def get_latest_file(directory):
    files = [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.xlsx') and 'KAM_NAMES' not in f]
    if not files:
        raise FileNotFoundError("No valid input files found.")
    return max(files, key=os.path.getmtime)

def process_gross_add(event_df):
    event_df['Reason'] = event_df['Reason'].str.strip().str.lower()
    included_reasons = ['port in', 'transfer ownership', 'n/a']
    event_df = event_df[event_df['Reason'].isin(included_reasons) | event_df['Reason'].isna()]

    event_df['Package Price'] = event_df['Amount'] + event_df['DiscountAmount']
    event_df['Slab'] = event_df['Package Price'].apply(calculate_slab)
    event_df['Commission (KD)'] = event_df['Package Price'] * event_df['Slab']

    columns_to_delete = ['OrderNumber', 'PriceType', 'PenaltyPeriod',
                         'Quantity', 'OfferType', 'MarketingCategory',
                         'TransactionDescription']
    event_df.drop(columns=[col for col in columns_to_delete if col in event_df.columns], inplace=True)

    total_commission = event_df['Commission (KD)'].sum()
    formatted_total = format_accounting(total_commission)

    total_row = pd.DataFrame([{col: "" for col in event_df.columns},
                              {'Commission (KD)': formatted_total}], index=['', 'Total'])
    event_df = pd.concat([event_df, total_row], ignore_index=True)

    return event_df

def process_upgrade(event_df):
    event_df = event_df.dropna(axis=1, how='all')
    event_df.insert(
        event_df.columns.get_loc('DiscountAmount') + 1,
        'New Package Price',
        event_df['Amount'] + event_df['DiscountAmount']
    )
    event_df.insert(
        event_df.columns.get_loc('OriginalDiscountAmount') + 1,
        'Package Price',
        event_df['OriginalAmount'] + event_df['OriginalDiscountAmount']
    )

    event_df['delta_kd'] = event_df['New Package Price'] - event_df['Package Price']
    event_df['Slab'] = event_df['delta_kd'].apply(calculate_slab)
    event_df['Commission (KD)'] = event_df['delta_kd'] * event_df['Slab']
    total_commission = event_df['Commission (KD)'].sum()
    formatted_total = format_accounting(total_commission)
    total_row = pd.DataFrame([{col: "" for col in event_df.columns},
                              {'Commission (KD)': formatted_total}], index=['', 'Total'])
    event_df = pd.concat([event_df, total_row], ignore_index=True)

    return event_df

def process_downgrade(event_df):
    # Delete all empty columns
    event_df = event_df.dropna(axis=1, how='all')

    # Add the 'six months completed' column with the formula
    event_df['six months completed'] = event_df['Age'].apply(
        lambda x: 'YES' if x > 180 else 'NO'
    )

    # Filter to keep only rows where 'six months completed' is 'NO'
    event_df = event_df[event_df['six months completed'] == 'NO']

    # Add 'New Package Price' column after 'Discount Amount'
    event_df.insert(
        event_df.columns.get_loc('DiscountAmount') + 1,
        'New Package Price',
        event_df['Amount'] + event_df['DiscountAmount']
    )
    event_df.insert(
        event_df.columns.get_loc('OriginalDiscountAmount') + 1,
        'Package Price',
        event_df['OriginalAmount'] + event_df['OriginalDiscountAmount']
    )
    event_df['delta_kd'] = event_df['New Package Price'] - event_df['Package Price']
    event_df['Slab'] = event_df['delta_kd'].apply(calculate_slab)
    event_df['Commission (KD)'] = -1 * event_df['delta_kd'] * event_df['Slab']
    total_commission = event_df['Commission (KD)'].sum()
    formatted_total = format_accounting(total_commission)
    total_row = pd.DataFrame([{col: "" for col in event_df.columns},
                              {'Commission (KD)': formatted_total}], index=['', 'Total'])
    event_df = pd.concat([event_df, total_row], ignore_index=True)

    return event_df

# ---- need to doublecheck not 100 % correct -----
def process_renewal(event_df):
    event_df = event_df.dropna(axis=1, how='all')
    event_df['service amount'] = event_df['Amount'] + event_df['DiscountAmount']
    event_df['Commission (KD)'] = event_df['service amount'] / 2
    total_commission = event_df['Commission (KD)'].sum()
    formatted_total = format_accounting(total_commission)
    total_row = pd.DataFrame([{col: "" for col in event_df.columns},
                              {'Commission (KD)': formatted_total}], index=['', 'Total'])
    event_df = pd.concat([event_df, total_row], ignore_index=True)

    return event_df



def process_file(file_path):
    shutil.copy(file_path, output_file_path)
    wb = load_workbook(output_file_path)

    first_sheet_name = wb.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=first_sheet_name)

    fonz_users = ['FONZ', 'FZ004360', 'FZ004361', 'V002741']
    df['New Payee ID'] = df['PayeeID'].apply(lambda x: 'FONZ' if x in fonz_users else x)
    df = df[df['New Payee ID'] == 'FONZ']

    event_types = ["ACTIVATION", "DEACTIVATION", "GROSS ADD", "CHURN",
                   "UPGRADE", "DOWNGRADE", "RENEW", "DEVICE ADD", "CLAWBACK"]

    for event_type in event_types:
        event_df = df[df['EventType'].str.upper() == event_type].copy()

        if event_df.empty:
            print(f"No data for {event_type}, skipping...")
            continue

        if event_type == "GROSS ADD":
            event_df = process_gross_add(event_df)
        elif event_type == "CHURN":
            event_df = process_churn(event_df)
        elif event_type == "UPGRADE":
            event_df = process_upgrade(event_df)
        elif event_type == "DOWNGRADE":
            event_df = process_downgrade(event_df)
        elif event_type == "RENEW": # need to doublecheck not 100 % correct
            event_df = process_renewal(event_df)

        if event_type not in wb.sheetnames:
            wb.create_sheet(event_type)
            print(f"Created new sheet for {event_type}")

        event_sheet = wb[event_type]

        for r_idx, row in enumerate(dataframe_to_rows(event_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                event_sheet.cell(row=r_idx, column=c_idx, value=value)

        tab = Table(
            displayName=event_type.replace(" ", ""),
            ref=f"A1:{get_column_letter(len(event_df.columns))}{len(event_df)+1}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tab.tableStyleInfo = style
        event_sheet.add_table(tab)

    wb.save(output_file_path)
    print(f"Processed and saved: {output_file_path}")
def main():
    try:
        latest_file = get_latest_file(INPUT_DIR)
        print(f"Processing latest file: {os.path.basename(latest_file)}")
        process_file(latest_file)
    except FileNotFoundError as e:
        print(str(e))

if __name__ == "__main__":
    main()
